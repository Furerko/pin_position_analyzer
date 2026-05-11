import sys
import os
import csv
import re
import math
from collections import Counter, defaultdict
from datetime import datetime
from statistics import median

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter


# ============================================================
# PIN POSITION ANALYZER V2
# Bardziej przejrzysty raport Excel + czytelne rzędy pinów
# ============================================================

ROBUST_Z_THRESHOLD = 5.0
MIN_VALID_VALUES_FOR_ANALYSIS = 10
IGNORE_ZERO_VALUES = True
ANALYZE_FAIL_ONLY = False
TOP_N = 15
TOP_BIGGEST_DEVIATIONS_PER_AXIS = 25

FALLBACK_TOLERANCE = {
    "X": 0.15,
    "Y": 0.15,
    "Z": 0.08,
}

GROSS_MIN = -1.0
GROSS_MAX = 25.0


# ============================================================
# MAPA RZĘDÓW PINÓW
# Tu ustawiasz fizyczny layout pinów.
# Program sam dobiera layout po nazwie Ref / nazwie pliku.
# ============================================================

PIN_LAYOUTS_BY_REF = {
    "GSR": {
        "ROW_1": [1, 5],
        "ROW_2": [2, 6],
        "ROW_3": [3, 7],
        "ROW_4": [4, 8],
    },

    # Dostosuj, jeśli SWEET ma inny layout fizyczny.
    # Domyślnie zostawione logicznie jako 1-5, 2-6, 3-7, 4-8, 9 osobno.
    "SWEET": {
        "ROW_1": [1, 5],
        "ROW_2": [2, 6],
        "ROW_3": [3, 7],
        "ROW_4": [4, 8],
        "ROW_5": [9],
    },

    # Dostosuj, jeśli HKMC ma inny layout fizyczny.
    # Domyślnie układ 16 pinów: 1-9, 2-10, ..., 8-16.
    "HKMC": {
        "ROW_1": [1, 9],
        "ROW_2": [2, 10],
        "ROW_3": [3, 11],
        "ROW_4": [4, 12],
        "ROW_5": [5, 13],
        "ROW_6": [6, 14],
        "ROW_7": [7, 15],
        "ROW_8": [8, 16],
    },
}

MIN_PINS_IN_ROW_FAIL = 2


# ============================================================
# HELPERS
# ============================================================

def get_exe_folder():
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def clean_cell(value):
    if value is None:
        return ""
    return str(value).strip().strip('"')


def parse_float(value):
    text = clean_cell(value).replace(",", ".")
    if text == "" or text.lower() in ["nan", "none", "null", "-"]:
        return None

    try:
        num = float(text)
        if math.isnan(num) or math.isinf(num):
            return None
        return num
    except Exception:
        return None


def parse_datetime(value):
    text = clean_cell(value)
    if not text:
        return None

    fmts = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%d.%m.%Y %H:%M:%S",
        "%d.%m.%Y %H:%M",
        "%m/%d/%Y, %I:%M:%S %p",
        "%m/%d/%Y %I:%M:%S %p",
        "%m/%d/%Y, %H:%M:%S",
        "%m/%d/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M:%S",
    ]

    for fmt in fmts:
        try:
            return datetime.strptime(text, fmt)
        except Exception:
            pass

    return None


def hour_bucket(dt):
    if dt is None:
        return "UNKNOWN_TIME"
    return dt.strftime("%Y-%m-%d %H:00")


def detect_dialect(file_path):
    try:
        with open(file_path, "r", encoding="utf-8", errors="ignore", newline="") as f:
            sample = f.read(12000)
            return csv.Sniffer().sniff(sample)
    except Exception:
        class SemiDialect(csv.excel):
            delimiter = ";"
        return SemiDialect


def pin_axis_from_metric(metric):
    m = re.search(r"Data[_ ]Pin(\d+)[_ ]Position([XYZ])", metric, re.IGNORECASE)
    if not m:
        return None, None
    return int(m.group(1)), m.group(2).upper()


def is_position_metric(metric):
    p, a = pin_axis_from_metric(metric)
    return p is not None and a is not None


def robust_stats(values):
    values = sorted([v for v in values if v is not None])
    if not values:
        return None, None

    med = median(values)
    mad = median([abs(v - med) for v in values])
    return med, mad


def robust_z(value, med, mad):
    if value is None or med is None or mad is None or mad == 0:
        return None
    return 0.6745 * (value - med) / mad


def style_header(ws):
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def auto_width(ws, max_width=60):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = 10

        for cell in col:
            if cell.value is not None:
                width = max(width, min(max_width, len(str(cell.value)) + 2))

        ws.column_dimensions[letter].width = width


def append_table(ws, headers, rows):
    ws.append(headers)

    for row in rows:
        ws.append(row)

    style_header(ws)
    auto_width(ws)


def add_bar_chart(ws, title, data_col, cat_col, end_row, pos):
    if end_row < 2:
        return

    chart = BarChart()
    chart.type = "bar"
    chart.title = title
    chart.y_axis.title = "Kategoria"
    chart.x_axis.title = "Liczba"

    data = Reference(ws, min_col=data_col, min_row=1, max_row=end_row)
    cats = Reference(ws, min_col=cat_col, min_row=2, max_row=end_row)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 16

    ws.add_chart(chart, pos)


def add_line_chart(ws, title, data_col, cat_col, end_row, pos):
    if end_row < 2:
        return

    chart = LineChart()
    chart.title = title
    chart.y_axis.title = "Liczba odchyleń"
    chart.x_axis.title = "Godzina"

    data = Reference(ws, min_col=data_col, min_row=1, max_row=end_row)
    cats = Reference(ws, min_col=cat_col, min_row=2, max_row=end_row)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 18

    ws.add_chart(chart, pos)


# ============================================================
# CSV PARSER
# ============================================================

def read_rows(file_path):
    dialect = detect_dialect(file_path)

    with open(file_path, "r", encoding="utf-8", errors="ignore", newline="") as f:
        return [row for row in csv.reader(f, dialect=dialect)]


def find_header(rows):
    for i, row in enumerate(rows):
        normalized = [clean_cell(c).upper() for c in row]

        if "NAME" in normalized:
            name_col = normalized.index("NAME")
            data_start = None

            for j, cell in enumerate(normalized):
                if cell == "USL (CALCULATED)":
                    data_start = j + 1
                    break

            if data_start is None:
                data_start = name_col + 16

            return i, name_col, data_start

    raise ValueError("Nie znaleziono wiersza z kolumną Name")


def collect_metrics(rows, header_idx, name_col):
    metrics = {}
    raw_rows = []

    for row in rows[header_idx + 1:]:
        name = clean_cell(row[name_col]) if len(row) > name_col else ""
        raw_rows.append((name, row))

        if name:
            metrics[name] = row

    return metrics, raw_rows


def find_datetime_row(raw_rows, data_start):
    best_row = None
    best_count = 0

    for name, row in raw_rows:
        count = 0

        for val in row[data_start:]:
            if parse_datetime(val):
                count += 1

        if count > best_count:
            best_count = count
            best_row = row

    return best_row if best_count > 0 else None


def detect_ref_from_file(file_base):
    upper = file_base.upper()

    for key in PIN_LAYOUTS_BY_REF.keys():
        if key in upper:
            return key

    return os.path.splitext(file_base)[0].split("_")[0].upper()


def extract_records(file_path):
    file_base = os.path.basename(file_path)

    rows = read_rows(file_path)
    header_idx, name_col, data_start = find_header(rows)

    header = rows[header_idx]
    metrics, raw_rows = collect_metrics(rows, header_idx, name_col)

    position_metrics = [m for m in metrics if is_position_metric(m)]

    if not position_metrics:
        raise ValueError("Nie znaleziono metryk Data_PinN_PositionX/Y/Z")

    max_len = max(len(row) for _, row in raw_rows) if raw_rows else len(header)

    dt_row = find_datetime_row(raw_rows, data_start)

    status_row = metrics.get("StatusBits") or metrics.get("Status_Bits")
    failure_row = metrics.get("Failure_Bits") or metrics.get("FailureBits")
    ref_row = metrics.get("Ref_Name") or metrics.get("RefName")

    records = []

    for col in range(data_start, max_len):
        sample_id = clean_cell(header[col]) if col < len(header) else ""

        dt = parse_datetime(dt_row[col]) if dt_row and col < len(dt_row) else None
        status = parse_float(status_row[col]) if status_row and col < len(status_row) else None
        failure = parse_float(failure_row[col]) if failure_row and col < len(failure_row) else None

        ref = clean_cell(ref_row[col]) if ref_row and col < len(ref_row) else ""
        ref = ref.upper() if ref else detect_ref_from_file(file_base)

        record = {
            "file": file_base,
            "ref": ref,
            "sample_index": col - data_start + 1,
            "sample_id": sample_id,
            "datetime": dt,
            "hour": hour_bucket(dt),
            "status_bits": status,
            "failure_bits": failure,
            "positions": {},
            "missing": {},
        }

        for metric in position_metrics:
            row = metrics[metric]
            pin, axis = pin_axis_from_metric(metric)

            raw = row[col] if col < len(row) else ""
            val = parse_float(raw)

            record["positions"][(pin, axis)] = val
            record["missing"][(pin, axis)] = clean_cell(raw) == "" or clean_cell(raw).lower() == "nan"

        records.append(record)

    return records


# ============================================================
# ANALIZA
# ============================================================

def is_fail_record(r):
    return (r.get("status_bits") not in [None, 0]) or (r.get("failure_bits") not in [None, 0])


def get_layout_for_ref(ref, active_pins):
    ref_upper = (ref or "").upper()

    for key, layout in PIN_LAYOUTS_BY_REF.items():
        if key in ref_upper:
            active = set(active_pins)

            return {
                row: [p for p in pins if p in active]
                for row, pins in layout.items()
                if any(p in active for p in pins)
            }

    # fallback auto
    pins = sorted(active_pins)
    rows = {}

    if not pins:
        return rows

    max_pin = max(pins)

    if max_pin <= 8:
        for i in range(1, 5):
            group = [p for p in [i, i + 4] if p in pins]
            if group:
                rows[f"ROW_{i}"] = group

    elif max_pin <= 16:
        for i in range(1, 9):
            group = [p for p in [i, i + 8] if p in pins]
            if group:
                rows[f"ROW_{i}"] = group

    else:
        for idx in range(0, len(pins), 4):
            rows[f"ROW_BLOCK_{idx // 4 + 1}"] = pins[idx:idx + 4]

    return rows


def build_baseline(records):
    values = defaultdict(list)
    quality = defaultdict(Counter)
    active_pins = defaultdict(set)

    for r in records:
        if ANALYZE_FAIL_ONLY and not is_fail_record(r):
            continue

        group = (r["file"], r["ref"])

        for (pin, axis), val in r["positions"].items():
            key = (r["file"], r["ref"], pin, axis)

            if r["missing"].get((pin, axis)):
                quality[key]["missing"] += 1
                continue

            if val is None:
                quality[key]["invalid"] += 1
                continue

            if val == 0:
                quality[key]["zero"] += 1

                if IGNORE_ZERO_VALUES:
                    continue

            if val < GROSS_MIN or val > GROSS_MAX:
                quality[key]["gross_excluded"] += 1
                continue

            values[key].append(val)
            quality[key]["valid"] += 1
            active_pins[group].add(pin)

    baseline = {}

    for key, vals in values.items():
        if len(vals) < MIN_VALID_VALUES_FOR_ANALYSIS:
            continue

        med, mad = robust_stats(vals)

        baseline[key] = {
            "n": len(vals),
            "median": med,
            "mad": mad,
            "mean": sum(vals) / len(vals),
            "min": min(vals),
            "max": max(vals),
        }

    return baseline, quality, active_pins


def classify_deviation(r, pin, axis, val, base):
    if val is None:
        return None

    if IGNORE_ZERO_VALUES and val == 0:
        return None

    if base is None:
        return None

    med = base["median"]
    mad = base["mad"]

    delta = val - med
    rz = robust_z(val, med, mad)

    reason = None

    if val < GROSS_MIN or val > GROSS_MAX:
        reason = "GROSS_OUTLIER"

    elif rz is not None and abs(rz) >= ROBUST_Z_THRESHOLD:
        reason = f"ROBUST_Z>{ROBUST_Z_THRESHOLD}"

    elif (mad is None or mad == 0) and abs(delta) >= FALLBACK_TOLERANCE.get(axis, 0.1):
        reason = f"ABS_DELTA>{FALLBACK_TOLERANCE.get(axis, 0.1)}"

    if not reason:
        return None

    return {
        "file": r["file"],
        "ref": r["ref"],
        "sample_index": r["sample_index"],
        "sample_id": r["sample_id"],
        "datetime": r["datetime"],
        "hour": r["hour"],
        "status_bits": r["status_bits"],
        "failure_bits": r["failure_bits"],
        "pin": pin,
        "axis": axis,
        "value": val,
        "median": med,
        "delta": delta,
        "abs_delta": abs(delta),
        "direction": "PLUS" if delta > 0 else "MINUS",
        "robust_z": rz,
        "reason": reason,
    }


def analyze(records):
    baseline, quality, active_pins = build_baseline(records)

    layouts = {
        group: get_layout_for_ref(group[1], pins)
        for group, pins in active_pins.items()
    }

    deviations = []
    row_events = []

    c_file = Counter()
    c_pin_axis = Counter()
    c_axis = Counter()
    c_direction = Counter()
    c_hour = Counter()
    c_file_hour = Counter()
    c_row_axis = Counter()

    for r in records:
        if ANALYZE_FAIL_ONLY and not is_fail_record(r):
            continue

        sample_devs = []

        for (pin, axis), val in r["positions"].items():
            key = (r["file"], r["ref"], pin, axis)

            dev = classify_deviation(r, pin, axis, val, baseline.get(key))

            if not dev:
                continue

            deviations.append(dev)
            sample_devs.append(dev)

            c_file[r["file"]] += 1
            c_pin_axis[(r["file"], r["ref"], pin, axis)] += 1
            c_axis[(r["file"], r["ref"], axis)] += 1
            c_direction[(r["file"], r["ref"], axis, dev["direction"])] += 1
            c_hour[r["hour"]] += 1
            c_file_hour[(r["file"], r["hour"])] += 1

        if sample_devs:
            group = (r["file"], r["ref"])

            existing = {
                (d["pin"], d["axis"]): d
                for d in sample_devs
            }

            for row_name, pins in layouts.get(group, {}).items():
                for axis in ["X", "Y", "Z"]:
                    failed = [
                        p for p in pins
                        if (p, axis) in existing
                    ]

                    if len(failed) >= MIN_PINS_IN_ROW_FAIL:
                        event = {
                            "file": r["file"],
                            "ref": r["ref"],
                            "sample_index": r["sample_index"],
                            "sample_id": r["sample_id"],
                            "datetime": r["datetime"],
                            "hour": r["hour"],
                            "row": row_name,
                            "axis": axis,
                            "pins": ",".join(str(p) for p in failed),
                            "count_pins": len(failed),
                        }

                        row_events.append(event)
                        c_row_axis[(r["file"], r["ref"], row_name, axis)] += 1

    return {
        "baseline": baseline,
        "quality": quality,
        "active_pins": active_pins,
        "layouts": layouts,
        "deviations": deviations,
        "row_events": row_events,
        "c_file": c_file,
        "c_pin_axis": c_pin_axis,
        "c_axis": c_axis,
        "c_direction": c_direction,
        "c_hour": c_hour,
        "c_file_hour": c_file_hour,
        "c_row_axis": c_row_axis,
    }


# ============================================================
# EXCEL REPORT V2
# ============================================================

def dt_text(dt):
    return dt.strftime("%Y-%m-%d %H:%M:%S") if dt else ""


def create_excel(output_file, analysis, records, input_files):
    wb = Workbook()

    ws = wb.active
    ws.title = "START_TUTAJ"

    devs = analysis["deviations"]
    row_events = analysis["row_events"]

    start_rows = [
        ["PIN POSITION REPORT V2", ""],
        ["Data raportu", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Liczba plików", len(input_files)],
        ["Liczba rekordów", len(records)],
        ["Liczba odchyłek", len(devs)],
        ["Problemy rzędów", len(row_events)],
        ["Najważniejsze zakładki", "1) Najwieksze_Odchylki  2) Os_X/Os_Y/Os_Z  3) Rzedy_Osie  4) Top_Pin_Os"],
        ["Metoda", "Mediana + MAD / Robust Z-score"],
        ["Próg Robust Z", ROBUST_Z_THRESHOLD],
    ]

    for row in start_rows:
        ws.append(row)

    ws[1][0].font = Font(bold=True, size=16)
    auto_width(ws)

    # ------------------------------------------------------------
    # Pliki
    # ------------------------------------------------------------

    ws_files = wb.create_sheet("Pliki")

    append_table(
        ws_files,
        ["Plik"],
        [[os.path.basename(f)] for f in input_files]
    )

    # ------------------------------------------------------------
    # Layout rzędów - bardzo czytelnie
    # ------------------------------------------------------------

    layout_rows = []

    for (file, ref), layout in sorted(analysis["layouts"].items()):
        for row_name, pins in layout.items():
            layout_rows.append([
                file,
                ref,
                row_name,
                ", ".join(f"Pin{p}" for p in pins)
            ])

    ws_layout = wb.create_sheet("Layout_Rzedow")

    append_table(
        ws_layout,
        ["File", "Ref", "Row", "Pins_in_row"],
        layout_rows
    )

    # ------------------------------------------------------------
    # Top Pin/Oś
    # ------------------------------------------------------------

    top_rows = []

    for (file, ref, pin, axis), cnt in analysis["c_pin_axis"].most_common():
        top_rows.append([
            file,
            ref,
            f"Pin{pin}",
            axis,
            cnt
        ])

    ws_top = wb.create_sheet("Top_Pin_Os")

    append_table(
        ws_top,
        ["File", "Ref", "Pin", "Axis", "Deviation_Count"],
        top_rows
    )

    add_bar_chart(
        ws_top,
        "TOP Pin/Oś - liczba odchyłek",
        5,
        3,
        min(ws_top.max_row, TOP_N + 1),
        "G2"
    )

    # ------------------------------------------------------------
    # Największe odchyłki z pliku
    # ------------------------------------------------------------

    biggest = sorted(
        devs,
        key=lambda x: (-x["abs_delta"], x["file"], x["axis"], x["pin"])
    )

    biggest_rows = []

    for e in biggest:
        biggest_rows.append([
            e["file"],
            e["ref"],
            e["axis"],
            f"Pin{e['pin']}",
            e["sample_index"],
            e["sample_id"],
            dt_text(e["datetime"]),
            e["hour"],
            e["value"],
            e["median"],
            e["delta"],
            e["abs_delta"],
            e["direction"],
            None if e["robust_z"] is None else round(e["robust_z"], 3),
            e["reason"],
        ])

    ws_big = wb.create_sheet("Najwieksze_Odchylki")

    append_table(
        ws_big,
        [
            "File",
            "Ref",
            "Axis",
            "Pin",
            "Sample_Index",
            "Sample_ID",
            "Datetime",
            "Hour",
            "Value",
            "Median",
            "Delta",
            "Abs_Delta",
            "Direction",
            "Robust_Z",
            "Reason",
        ],
        biggest_rows
    )

    if ws_big.max_row > 1:
        ws_big.conditional_formatting.add(
            f"L2:L{ws_big.max_row}",
            ColorScaleRule(
                start_type="min",
                start_color="FFFFFF",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="max",
                end_color="F8696B",
            )
        )

    # ------------------------------------------------------------
    # Osobne zakładki dla osi X/Y/Z z największymi odchyłkami
    # ------------------------------------------------------------

    for axis in ["X", "Y", "Z"]:
        axis_devs = [e for e in biggest if e["axis"] == axis]

        rows = []

        for e in axis_devs[:TOP_BIGGEST_DEVIATIONS_PER_AXIS * max(1, len(input_files))]:
            rows.append([
                e["file"],
                e["ref"],
                f"Pin{e['pin']}",
                e["sample_index"],
                e["sample_id"],
                dt_text(e["datetime"]),
                e["value"],
                e["median"],
                e["delta"],
                e["abs_delta"],
                e["direction"],
                None if e["robust_z"] is None else round(e["robust_z"], 3),
                e["reason"],
            ])

        ws_axis = wb.create_sheet(f"Os_{axis}")

        append_table(
            ws_axis,
            [
                "File",
                "Ref",
                "Pin",
                "Sample_Index",
                "Sample_ID",
                "Datetime",
                "Value",
                "Median",
                "Delta",
                "Abs_Delta",
                "Direction",
                "Robust_Z",
                "Reason",
            ],
            rows
        )

        if ws_axis.max_row > 1:
            ws_axis.conditional_formatting.add(
                f"J2:J{ws_axis.max_row}",
                ColorScaleRule(
                    start_type="min",
                    start_color="FFFFFF",
                    mid_type="percentile",
                    mid_value=50,
                    mid_color="FFEB84",
                    end_type="max",
                    end_color="F8696B",
                )
            )

            add_bar_chart(
                ws_axis,
                f"Największe odchyłki - oś {axis}",
                10,
                3,
                min(ws_axis.max_row, TOP_N + 1),
                "O2"
            )

    # ------------------------------------------------------------
    # Rzędy wg osi
    # ------------------------------------------------------------

    row_axis_rows = []

    for (file, ref, row_name, axis), cnt in analysis["c_row_axis"].most_common():
        row_axis_rows.append([
            file,
            ref,
            row_name,
            axis,
            cnt
        ])

    ws_ra = wb.create_sheet("Rzedy_Osie")

    append_table(
        ws_ra,
        ["File", "Ref", "Row", "Axis", "Row_Axis_Deviation_Count"],
        row_axis_rows
    )

    add_bar_chart(
        ws_ra,
        "Problemy rzędów wg osi",
        5,
        3,
        min(ws_ra.max_row, TOP_N + 1),
        "G2"
    )

    # ------------------------------------------------------------
    # Szczegóły problemów rzędów
    # ------------------------------------------------------------

    re_rows = []

    for e in row_events:
        re_rows.append([
            e["file"],
            e["ref"],
            e["row"],
            e["axis"],
            e["pins"],
            e["count_pins"],
            e["sample_index"],
            e["sample_id"],
            dt_text(e["datetime"]),
            e["hour"],
        ])

    ws_re = wb.create_sheet("Rzedy_Szczegoly")

    append_table(
        ws_re,
        [
            "File",
            "Ref",
            "Row",
            "Axis",
            "Pins",
            "Count_Pins",
            "Sample_Index",
            "Sample_ID",
            "Datetime",
            "Hour",
        ],
        re_rows
    )

    # ------------------------------------------------------------
    # Trend godzinowy
    # ------------------------------------------------------------

    tr_rows = []

    for (file, hour), cnt in sorted(
        analysis["c_file_hour"].items(),
        key=lambda x: (x[0][0], x[0][1])
    ):
        tr_rows.append([
            file,
            hour,
            cnt
        ])

    ws_tr = wb.create_sheet("Trend_Godzina")

    append_table(
        ws_tr,
        ["File", "Hour", "Deviation_Count"],
        tr_rows
    )

    add_line_chart(
        ws_tr,
        "Trend godzinowy odchyłek",
        3,
        2,
        ws_tr.max_row,
        "E2"
    )

    # ------------------------------------------------------------
    # Kierunek osi
    # ------------------------------------------------------------

    dir_rows = []

    for (file, ref, axis, direction), cnt in analysis["c_direction"].most_common():
        dir_rows.append([
            file,
            ref,
            axis,
            direction,
            cnt
        ])

    ws_dir = wb.create_sheet("Kierunek_Osi")

    append_table(
        ws_dir,
        ["File", "Ref", "Axis", "Direction", "Count"],
        dir_rows
    )

    add_bar_chart(
        ws_dir,
        "Kierunek odchyłek",
        5,
        3,
        min(ws_dir.max_row, TOP_N + 1),
        "G2"
    )

    # ------------------------------------------------------------
    # Baseline
    # ------------------------------------------------------------

    bl_rows = []

    for (file, ref, pin, axis), st in sorted(analysis["baseline"].items()):
        bl_rows.append([
            file,
            ref,
            f"Pin{pin}",
            axis,
            st["n"],
            st["median"],
            st["mad"],
            st["mean"],
            st["min"],
            st["max"],
        ])

    ws_bl = wb.create_sheet("Baseline")

    append_table(
        ws_bl,
        [
            "File",
            "Ref",
            "Pin",
            "Axis",
            "N",
            "Median",
            "MAD",
            "Mean",
            "Min",
            "Max",
        ],
        bl_rows
    )

    # ------------------------------------------------------------
    # Jakość danych
    # ------------------------------------------------------------

    q_rows = []

    for (file, ref, pin, axis), q in sorted(analysis["quality"].items()):
        q_rows.append([
            file,
            ref,
            f"Pin{pin}",
            axis,
            q.get("valid", 0),
            q.get("zero", 0),
            q.get("missing", 0),
            q.get("invalid", 0),
            q.get("gross_excluded", 0),
        ])

    ws_q = wb.create_sheet("Jakosc_Danych")

    append_table(
        ws_q,
        [
            "File",
            "Ref",
            "Pin",
            "Axis",
            "Valid",
            "Zero",
            "Missing",
            "Invalid",
            "Gross_Excluded",
        ],
        q_rows
    )

    # ------------------------------------------------------------
    # Wszystkie odchyłki
    # ------------------------------------------------------------

    dev_rows = []

    for e in devs:
        dev_rows.append([
            e["file"],
            e["ref"],
            f"Pin{e['pin']}",
            e["axis"],
            e["sample_index"],
            e["sample_id"],
            dt_text(e["datetime"]),
            e["hour"],
            e["status_bits"],
            e["failure_bits"],
            e["value"],
            e["median"],
            e["delta"],
            e["abs_delta"],
            e["direction"],
            None if e["robust_z"] is None else round(e["robust_z"], 3),
            e["reason"],
        ])

    ws_dev = wb.create_sheet("Wszystkie_Odchylki")

    append_table(
        ws_dev,
        [
            "File",
            "Ref",
            "Pin",
            "Axis",
            "Sample_Index",
            "Sample_ID",
            "Datetime",
            "Hour",
            "StatusBits",
            "Failure_Bits",
            "Value",
            "Median",
            "Delta",
            "Abs_Delta",
            "Direction",
            "Robust_Z",
            "Reason",
        ],
        dev_rows
    )

    wb.save(output_file)


# ============================================================
# CONSOLE
# ============================================================

def print_summary(analysis, records):
    print("\n" + "=" * 100)
    print("PIN POSITION ANALYZER V2 - PODSUMOWANIE")
    print("=" * 100)

    print(f"Rekordy: {len(records)}")
    print(f"Odchyłki: {len(analysis['deviations'])}")
    print(f"Problemy rzędów: {len(analysis['row_events'])}")

    print("\nMAPA RZĘDÓW UŻYTA W ANALIZIE:")

    for (file, ref), layout in sorted(analysis["layouts"].items()):
        print(f"\n{file} | REF: {ref}")

        for row, pins in layout.items():
            print(f"  {row}: " + ", ".join(f"Pin{p}" for p in pins))

    print("\nTOP PIN/OŚ:")

    for (file, ref, pin, axis), cnt in analysis["c_pin_axis"].most_common(TOP_N):
        print(f"  {file} | {ref} | Pin{pin} | {axis}: {cnt}")

    print("\nNAJWIĘKSZE ODCHYŁKI:")

    biggest = sorted(
        analysis["deviations"],
        key=lambda x: -x["abs_delta"]
    )[:TOP_N]

    for e in biggest:
        print(
            f"  {e['file']} | {e['ref']} | "
            f"Pin{e['pin']} {e['axis']} | "
            f"Value={e['value']} | "
            f"Median={e['median']:.4f} | "
            f"Delta={e['delta']:.4f}"
        )

    print("=" * 100)


# ============================================================
# MAIN
# ============================================================

def main():
    if len(sys.argv) < 2:
        print("Przeciągnij jeden lub więcej plików CSV na program EXE.")
        input("ENTER aby wyjść...")
        return

    csv_files = [
        p for p in sys.argv[1:]
        if p.lower().endswith(".csv") and os.path.isfile(p)
    ]

    if not csv_files:
        print("Nie znaleziono plików CSV.")
        input("ENTER aby wyjść...")
        return

    all_records = []

    print("Analizowane pliki:")

    for f in csv_files:
        print("-", f)

        try:
            recs = extract_records(f)
            all_records.extend(recs)
            print(f"  OK: {len(recs)} rekordów")

        except Exception as e:
            print(f"  BŁĄD: {e}")

    if not all_records:
        print("Brak danych do analizy.")
        input("ENTER aby wyjść...")
        return

    analysis = analyze(all_records)

    print_summary(analysis, all_records)

    out_dir = get_exe_folder()
    timestamp = datetime.now().strftime("%Y_%m_%d_%H_%M_%S")

    output = os.path.join(
        out_dir,
        f"pin_position_report_V2__{timestamp}.xlsx"
    )

    create_excel(
        output,
        analysis,
        all_records,
        csv_files
    )

    print("\nZapisano raport Excel:")
    print(output)

    input("\nENTER aby zamknąć...")


if __name__ == "__main__":
    main()
